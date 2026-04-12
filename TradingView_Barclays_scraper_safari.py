!pip install selenium pandas webdriver_manager
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import re
import time

URL = "https://www.tradingview.com/symbols/LSE-BARC/forecast/"

# Safari is built into macOS — no driver download needed
driver = webdriver.Safari()
driver.set_window_size(1280, 900)

try:
    driver.get(URL)

    # TradingView is JS-heavy; give it time to render fully
    time.sleep(6)

    valor = None

    # Strategy 1: partial class match on 'priceWrapper' (hash-resistant)
    candidates = driver.find_elements(By.CSS_SELECTOR, "[class*='priceWrapper']")
    for el in candidates:
        text = el.text.strip().split('\n')[0]
        if re.match(r'^\d{2,5}(\.\d+)?$', text):
            valor = text
            break

    # Strategy 2: broader sweep for price-like numbers near the forecast section
    if not valor:
        candidates = driver.find_elements(
            By.XPATH,
            "//*[contains(@class,'price') or contains(@class,'Price')]"
        )
        for el in candidates:
            text = el.text.strip().split('\n')[0]
            if re.match(r'^\d{2,5}(\.\d+)?$', text):
                valor = text
                break

    if valor:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Price Targets"
        ws["A1"], ws["B1"] = "Ticker", "Price Target (GBX)"
        ws["A2"] = "BARC"
        ws["B2"] = float(valor)
        wb.save("price_target.xlsx")
        print(f"Saved: {valor} → price_target.xlsx")
    else:
        print("Price target not found. TradingView may have blocked the request or changed its layout.")

except Exception as e:
    print(f"Error: {e}")

finally:
    driver.quit()
